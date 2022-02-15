# ! python3
import re
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Union, Set

import requests
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl import Workbook, load_workbook

# general information about regex and the specific features used below
# https://docs.python.org/3/howto/regex.html#non-capturing-and-named-groups
# https://docs.python.org/3/howto/regex.html#greedy-versus-non-greedy

# use link below to see how this regex works:
# https://pythex.org/?regex=%5C((%3FP%3Cinner%3E.*%3F)%5C)&test_string=%20%20%20asdf%20%20%20(%20%20location%20)%20asdf&ignorecase=1&multiline=0&dotall=0&verbose=0
LOCATION_REGEX = re.compile('\((?P<inner>.*?)\)')

# use link below to see how this regex works:
# https://pythex.org/?regex=(%3FP%3Cbr%3E%5Cd%2B)br&test_string=%242%2C064%2F%203br%20-%201362ft2%20-%20&ignorecase=1&multiline=0&dotall=0&verbose=0
BED_REGEX = re.compile('(?P<beds>\d+)br', re.IGNORECASE)


def process_result(result_element: Tag):
    date_element: Tag = result_element.select_one('time.result-date')
    posted_datetime: datetime = datetime.strptime(date_element['datetime'], '%Y-%m-%d %H:%M')

    detail_element: Tag = result_element.select_one('span.housing')
    # https://realpython.com/python-walrus-operator/
    if detail_element is not None and (match := BED_REGEX.search(detail_element.text)):
        beds = int(match.group('beds'))
    else:
        beds = 0

    link: Tag = result_element.select_one('h3 a')
    price: Tag = result_element.select_one('span.result-price')

    location_element: Tag = result_element.select_one('span.result-hood')
    location = LOCATION_REGEX.search(str(location_element)).group('inner').strip()

    return {
        'name': link.text,
        'beds': beds,
        'date': posted_datetime.date(),
        'link': link['href'],
        'price': int(price.text.replace('$', '').replace(',', '')),
        'location': location
    }


def result_gen(url: str, max_results: int = 500, min_beds: int = None) -> Iterable[Dict[str, str]]:
    # https://realpython.com/introduction-to-python-generators/
    soup = BeautifulSoup(requests.get(url).content, 'lxml')

    i = 0  # used to keep track of how many results have been returned
    while True:
        # see the below links for information about the CSS selectors used in soup.select(...)
        # https://www.w3schools.com/csSref/sel_id.asp
        # https://www.w3schools.com/cssref/sel_class.asp
        for res in soup.select('#search-results li.result-row div.result-info'):
            if i > max_results:
                break  # break the for loop if more than the max number of results have been returned
            else:
                result = process_result(result_element=res)
                if min_beds is not None:
                    if result['beds'] < min_beds:
                        continue

                i += 1
                yield result

        if i > max_results:
            break  # might also need to break the while loop

        try:  # try to keep getting the next page, and break the while loop if any errors happen
            # https://www.w3schools.com/cssref/sel_attr_begin.asp
            next_page_link = soup.select_one('a[title^=next]')

            # generate the URL for the next page by replacing the path portion with that from the href attribute
            next_url = urllib.parse.urlunsplit(
                # https://docs.python.org/3/library/urllib.parse.html#urllib.parse.urlsplit
                urllib.parse.urlsplit(url)._replace(path=next_page_link['href'])
            )

            print(f'Getting next page: {next_url}, {i} results so far')
            soup = BeautifulSoup(requests.get(next_url).content, 'lxml')
        except Exception as e:
            break


def save_results(path: Union[str, Path], results: Iterable[Dict]):
    # make sure that the path variable is a Path object, which is a good practice and makes lots of operations easier
    path = Path(path) if not isinstance(path, Path) else path

    # load the workbook if the designated file already exists, otherwise generate a blank Workbook in memory
    book = load_workbook(filename=str(path)) if path.exists() else Workbook()
    sheet = book.worksheets[0]

    # set up the header row and resize the columns if there's no data present
    if sheet.max_row <= 1:
        sheet.column_dimensions['A'].width = 65
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 10

        headers = [
            'Item Name',
            'Beds',
            'Price',
            'Link',
            'Location',
            'Posted Date'
        ]

        for i, h in enumerate(headers):
            sheet.cell(row=1, column=i + 1).value = h

    for i, res in enumerate(results):
        print(f'Saving ${res["price"]:,} - {res["name"]}')
        first_free_row = sheet.max_row + 1
        sheet.cell(row=first_free_row, column=1).value = res['name']

        sheet.cell(row=first_free_row, column=2).value = res['beds']

        sheet.cell(row=first_free_row, column=3).value = res['price']
        sheet.cell(row=first_free_row, column=3).number_format = '$#,##0'

        sheet.cell(row=first_free_row, column=4).value = f"=HYPERLINK(\"{res['link']}\")"
        sheet.cell(row=first_free_row, column=4).style = 'Hyperlink'

        sheet.cell(row=first_free_row, column=5).value = res['location']

        sheet.cell(row=first_free_row, column=6).value = res['date']
        sheet.cell(row=first_free_row, column=6).number_format = 'd-mmm;@'

    book.save(filename=str(path))


def load_previous_results(path: Union[str, Path]) -> Set:
    if path.exists():
        prev_results = load_workbook(
            filename=str(path),
            read_only=True,
            data_only=True
        ).worksheets[0]

        return set(
            val[0]  # iter_rows returns a tuple with a single value
            # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.iter_rows
            for val in prev_results.iter_rows(
                min_row=2, max_row=prev_results.max_row,
                min_col=3, max_col=3,
                values_only=True
            )
        )
    else:
        return set()


# if the date of the posting is todays date it will notify your desktop
# if str(res['date']) == today.strftime('%b %d'):
#     # https://github.com/ms7m/notify-py#usage
#     notification = Notify()
#     notification.title = res['name']
#     notification.message = f"{res['link']}\n{res['price']}"
#     notification.send()

# https://docs.python.org/3/library/__main__.html#idiomatic-usage
# https://stackoverflow.com/questions/419163/what-does-if-name-main-do
if __name__ == '__main__':
    # needs to have a urls.txt file in the same folder. The file should have 1 URL per line
    URL_LIST = Path('urls.txt').open('r').readlines()

    RESULT_PATH = Path('Apartments.xlsx')

    prev_results = load_previous_results(RESULT_PATH)
    results = (
        result
        for url in URL_LIST
        for result in result_gen(url, max_results=500, min_beds=3)
        if result['link'] not in prev_results
    )

    save_results(path=RESULT_PATH, results=results)
